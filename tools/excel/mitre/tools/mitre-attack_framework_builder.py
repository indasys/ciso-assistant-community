#!/usr/bin/env python3
"""Build the CISO Assistant MITRE ATT&CK Excel library.

The complete pipeline is handled by this script:

1. Download README.md, LICENSE.txt and enterprise-attack.json from mitre/cti.
2. Read the ATT&CK version from the JSON file's latest release commit.
3. Export the legacy intermediate techniques/measures workbooks.
4. Build the final, versioned CISO Assistant workbook.
5. Delete downloaded and intermediate files unless `-k/--keep` is used.

Run from shell:

    python ./mitre-attack_framework_builder.py
    python ./mitre-attack_framework_builder.py --keep

Manual steps for French translations in the final workbook:

- To activate the French translation formulas, manually remove the `@` in
  front of the `=`. Excel's Find and Replace tool can help, but process about
  100 cells at a time instead of replacing every cell at once to avoid hitting
  the translation API limit too quickly.
- After the formulas finish translating, copy the translated cells and paste
  them back into the same location using Paste Values. This ensures that the
  exported YAML contains the actual translations; otherwise, it may contain
  unexpected formula-related text instead.
- Some source text may be too long for the translation function. Translate
  those cells manually (e.g. with DeepL).
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence
from urllib.parse import urlparse

import requests
from mitreattack.stix20 import MitreAttackData
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet

try:
    from tqdm import tqdm
except ModuleNotFoundError:  # pragma: no cover
    tqdm = None


# ---------------------------------------------------------------------------
# Fixed configuration
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent

RAW_BASE_URL = "https://raw.githubusercontent.com/mitre/cti/master"
README_URL = f"{RAW_BASE_URL}/README.md"
LICENSE_URL = f"{RAW_BASE_URL}/LICENSE.txt"
ENTERPRISE_ATTACK_URL = f"{RAW_BASE_URL}/enterprise-attack/enterprise-attack.json"
COMMITS_API_URL = "https://api.github.com/repos/mitre/cti/commits"
SOURCE_URL = "https://github.com/mitre/cti"

README_PATH = SCRIPT_DIR / "README.md"
LICENSE_PATH = SCRIPT_DIR / "LICENSE.txt"
ENTERPRISE_ATTACK_PATH = SCRIPT_DIR / "enterprise-attack.json"
TECHNIQUES_PATH = SCRIPT_DIR / "techniques.xlsx"
MEASURES_PATH = SCRIPT_DIR / "measures.xlsx"

LIBRARY_URN = "urn:intuitem:risk:library:mitre-attack"
THREATS_BASE_URN = "urn:intuitem:risk:threat:mitre-attack"
MITIGATIONS_BASE_URN = "urn:intuitem:risk:function:mitre-attack"
LIBRARY_VERSION = "1"
LIBRARY_LOCALE = "en"
LIBRARY_REF_ID = "mitre-attack"
LIBRARY_PROVIDER = "Mitre ATT&CK"
LIBRARY_PACKAGER = "intuitem"

LIBRARY_META_SHEET = "library_meta"
THREATS_META_SHEET = "threats_meta"
THREATS_CONTENT_SHEET = "threats_content"
MITIGATIONS_META_SHEET = "mitigations_meta"
MITIGATIONS_CONTENT_SHEET = "mitigations_content"

VERSION_PATTERN = re.compile(
    r"ATT&CK\s+v(?P<version>\d+(?:\.\d+)*)\s+Enterprise\b",
    flags=re.IGNORECASE,
)

# The source STIX objects do not provide a NIST CSF function. These values are
# preserved from the existing mitre-attack.xlsx library. A future mitigation
# which is not in this mapping is assigned to "protect" and reported.
MITIGATION_CSF_FUNCTIONS = {
    "M1013": "govern",
    "M1015": "protect",
    "M1016": "detect",
    "M1017": "govern",
    "M1018": "protect",
    "M1019": "protect",
    "M1020": "protect",
    "M1021": "protect",
    "M1022": "protect",
    "M1024": "protect",
    "M1025": "protect",
    "M1026": "protect",
    "M1027": "protect",
    "M1028": "protect",
    "M1029": "protect",
    "M1030": "protect",
    "M1031": "detect",
    "M1032": "protect",
    "M1033": "protect",
    "M1034": "protect",
    "M1035": "protect",
    "M1036": "protect",
    "M1037": "protect",
    "M1038": "protect",
    "M1039": "protect",
    "M1040": "detect",
    "M1041": "protect",
    "M1042": "protect",
    "M1043": "protect",
    "M1044": "protect",
    "M1045": "protect",
    "M1046": "protect",
    "M1047": "protect",
    "M1048": "protect",
    "M1049": "detect",
    "M1050": "detect",
    "M1051": "protect",
    "M1052": "protect",
    "M1053": "recover",
    "M1054": "protect",
    "M1055": "govern",
    "M1056": "govern",
    "M1057": "detect",
    "M1060": "protect",
}


@dataclass(frozen=True)
class AttackRecord:
    ref_id: str
    name: str
    description: str


# ---------------------------------------------------------------------------
# Display and download helpers
# ---------------------------------------------------------------------------


def display_path(path: Path) -> str:
    try:
        return str(path.relative_to(SCRIPT_DIR))
    except ValueError:
        return str(path)


def print_step_banner(step_number: int, title: str) -> None:
    message = f"##### [STEP {step_number}] {title} #####"
    line = "#" * len(message)
    print(f"\n{line}\n{message}\n{line}\n")


def validate_download_url(url: str) -> None:
    parsed = urlparse(url)
    if parsed.scheme != "https":
        raise ValueError(f"Unsupported URL scheme: {parsed.scheme!r}")
    if parsed.hostname not in {"raw.githubusercontent.com", "api.github.com"}:
        raise ValueError(f"Unexpected download host: {parsed.hostname!r}")


def create_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "Accept": "application/vnd.github+json",
            "Accept-Encoding": "identity",
            "User-Agent": "ciso-assistant-mitre-attack-framework-builder",
        }
    )
    return session


def download_file(
    session: requests.Session,
    url: str,
    destination: Path,
) -> None:
    validate_download_url(url)
    temporary_path = destination.with_name(f"{destination.name}.part")
    temporary_path.unlink(missing_ok=True)

    print(f'📥 [DOWN] Downloading: "{url}"')
    try:
        with session.get(url, stream=True, timeout=60) as response:
            response.raise_for_status()
            total_size = int(response.headers.get("Content-Length", 0))
            with temporary_path.open("wb") as output:
                chunks = response.iter_content(chunk_size=64 * 1024)
                if tqdm is None:
                    for chunk in chunks:
                        if chunk:
                            output.write(chunk)
                else:
                    with tqdm(
                        total=total_size if total_size > 0 else None,
                        unit="o",
                        unit_scale=True,
                        unit_divisor=1024,
                        desc=destination.name,
                    ) as progress:
                        for chunk in chunks:
                            if not chunk:
                                continue
                            output.write(chunk)
                            progress.update(len(chunk))
        temporary_path.replace(destination)
    finally:
        temporary_path.unlink(missing_ok=True)

    print(f'✅ [OK] Downloaded: "{display_path(destination)}"')


def download_sources(session: requests.Session) -> list[Path]:
    downloaded: list[Path] = []
    try:
        for url, destination in (
            (README_URL, README_PATH),
            (LICENSE_URL, LICENSE_PATH),
            (ENTERPRISE_ATTACK_URL, ENTERPRISE_ATTACK_PATH),
        ):
            download_file(session, url, destination)
            downloaded.append(destination)
    except Exception:
        cleanup(downloaded)
        raise
    return downloaded


def get_attack_version(session: requests.Session) -> str:
    validate_download_url(COMMITS_API_URL)
    response = session.get(
        COMMITS_API_URL,
        params={
            "path": "enterprise-attack/enterprise-attack.json",
            "sha": "master",
            "per_page": 100,
        },
        timeout=60,
    )
    response.raise_for_status()

    commits = response.json()
    if not isinstance(commits, list):
        raise ValueError("Unexpected response from the GitHub commits API")

    for item in commits:
        message = item.get("commit", {}).get("message", "")
        match = VERSION_PATTERN.search(message)
        if match:
            return match.group("version")

    raise ValueError(
        "Unable to find an 'ATT&CK v# Enterprise' commit for enterprise-attack.json"
    )


# ---------------------------------------------------------------------------
# Source metadata extraction
# ---------------------------------------------------------------------------


def extract_markdown_section(markdown: str, heading: str) -> str:
    lines = markdown.splitlines()
    expected_heading = f"## {heading}"

    try:
        start = next(
            index
            for index, line in enumerate(lines)
            if line.strip() == expected_heading
        )
    except StopIteration as exc:
        raise ValueError(
            f"Section {expected_heading!r} not found in README.md"
        ) from exc

    end = next(
        (
            index
            for index in range(start + 1, len(lines))
            if lines[index].startswith("## ")
        ),
        len(lines),
    )
    section = "\n".join(lines[start + 1 : end]).strip()
    if not section:
        raise ValueError(f"Section {expected_heading!r} is empty in README.md")
    return section


def build_library_description(readme: str) -> str:
    attack_section = extract_markdown_section(readme, "ATT&CK")
    attack_section = re.sub(r"<(https?://[^>]+)>", r"\1", attack_section)
    return f"{attack_section}\n\nSource: {SOURCE_URL}"


def extract_attack_license(license_text: str) -> str:
    lines = license_text.splitlines()

    try:
        start = next(
            index for index, line in enumerate(lines) if line.strip() == "ATT&CK®"
        )
        end = next(
            index
            for index in range(start + 1, len(lines))
            if lines[index].strip() == "CAPEC™"
        )
    except StopIteration as exc:
        raise ValueError("ATT&CK license section not found in LICENSE.txt") from exc

    section = lines[start + 1 : end]
    while section and (not section[0].strip() or set(section[0].strip()) == {"="}):
        section.pop(0)
    while section and not section[-1].strip():
        section.pop()

    formatted: list[str] = []
    index = 0
    while index < len(section):
        line = section[index]
        stripped = line.strip()
        next_is_underline = (
            index + 1 < len(section)
            and bool(section[index + 1].strip())
            and set(section[index + 1].strip()) == {"-"}
        )

        if stripped in {"License", "Disclaimers"} and next_is_underline:
            if formatted and formatted[-1] != "":
                formatted.append("")
            formatted.extend((f"## {stripped}", ""))
            index += 2
            continue

        formatted.append(line.rstrip())
        index += 1

    copyright_text = "\n".join(formatted).strip()
    if not copyright_text.startswith("## License\n\n"):
        raise ValueError("The ATT&CK License heading could not be formatted")
    if "\n## Disclaimers\n\n" not in copyright_text:
        raise ValueError("The ATT&CK Disclaimers heading could not be formatted")
    return copyright_text


# ---------------------------------------------------------------------------
# ATT&CK extraction
# ---------------------------------------------------------------------------


def get_external_reference(stix_object: object) -> tuple[str, str]:
    references = getattr(stix_object, "external_references", ())
    for reference in references:
        if getattr(reference, "source_name", None) != "mitre-attack":
            continue
        ref_id = getattr(reference, "external_id", None)
        url = getattr(reference, "url", None)
        if ref_id and url:
            return str(ref_id), str(url)

    raise ValueError(
        f"No complete mitre-attack external reference for "
        f"{getattr(stix_object, 'id', '<unknown>')}"
    )


def to_attack_record(stix_object: object) -> AttackRecord:
    ref_id, url = get_external_reference(stix_object)
    name = str(getattr(stix_object, "name")).strip()
    description = str(getattr(stix_object, "description", "")).strip()
    if not description:
        description = url
    elif url not in description:
        description = f"{description}\n{url}"
    return AttackRecord(ref_id=ref_id, name=name, description=description)


def extract_attack_records(
    source: Path,
) -> tuple[list[AttackRecord], list[AttackRecord]]:
    attack_data = MitreAttackData(str(source))
    techniques = sorted(
        (
            to_attack_record(item)
            for item in attack_data.get_techniques(remove_revoked_deprecated=True)
        ),
        key=lambda item: item.ref_id,
    )
    mitigations = sorted(
        (
            to_attack_record(item)
            for item in attack_data.get_mitigations(remove_revoked_deprecated=True)
        ),
        key=lambda item: item.ref_id,
    )
    return techniques, mitigations


# ---------------------------------------------------------------------------
# Workbook generation
# ---------------------------------------------------------------------------


def style_meta_sheet(worksheet: Worksheet) -> None:
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 120


def style_content_sheet(
    worksheet: Worksheet,
    widths: Sequence[float],
) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column_index, width in enumerate(widths, start=1):
        column_letter = worksheet.cell(row=1, column=column_index).column_letter
        worksheet.column_dimensions[column_letter].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def write_simple_workbook(
    output: Path,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    widths: Sequence[float],
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    style_content_sheet(worksheet, widths)
    workbook.save(output)
    workbook.close()


def write_intermediate_workbooks(
    techniques: Sequence[AttackRecord],
    mitigations: Sequence[AttackRecord],
) -> list[Path]:
    generated: list[Path] = []
    try:
        write_simple_workbook(
            TECHNIQUES_PATH,
            ("ref_id", "name", "description"),
            ((item.ref_id, item.name, item.description) for item in techniques),
            (14, 48, 100),
        )
        generated.append(TECHNIQUES_PATH)
        write_simple_workbook(
            MEASURES_PATH,
            ("ref_id", "name", "category", "description"),
            (
                (item.ref_id, item.name, "technical", item.description)
                for item in mitigations
            ),
            (14, 48, 18, 100),
        )
        generated.append(MEASURES_PATH)
    except Exception:
        cleanup(generated)
        raise
    return generated


def append_meta_rows(
    worksheet: Worksheet,
    rows: Iterable[tuple[str, object]],
) -> None:
    for key, value in rows:
        worksheet.append((key, value))
    style_meta_sheet(worksheet)


def build_final_workbook(
    output: Path,
    version: str,
    description: str,
    copyright_text: str,
    techniques: Sequence[AttackRecord],
    mitigations: Sequence[AttackRecord],
) -> None:
    workbook = Workbook()
    library_meta = workbook.active
    library_meta.title = LIBRARY_META_SHEET
    threats_meta = workbook.create_sheet(THREATS_META_SHEET)
    threats_content = workbook.create_sheet(THREATS_CONTENT_SHEET)
    mitigations_meta = workbook.create_sheet(MITIGATIONS_META_SHEET)
    mitigations_content = workbook.create_sheet(MITIGATIONS_CONTENT_SHEET)

    library_name = f"Mitre ATT&CK v{version} - Threats and Mitigations"
    # See the module docstring for the manual steps required to activate the
    # French translation formulas and replace them with their resulting values.
    append_meta_rows(
        library_meta,
        (
            ("type", "library"),
            ("urn", LIBRARY_URN),
            ("version", LIBRARY_VERSION),
            ("locale", LIBRARY_LOCALE),
            ("ref_id", LIBRARY_REF_ID),
            ("name", library_name),
            ("description", description),
            ("copyright", copyright_text),
            ("provider", LIBRARY_PROVIDER),
            ("packager", LIBRARY_PACKAGER),
            ("name[fr]", "=TRADUIRE(B6,\"en\",\"fr\")"),
            ("description[fr]", "=TRADUIRE(B7,\"en\",\"fr\")"),
        ),
    )

    append_meta_rows(
        threats_meta,
        (("type", "threats"), ("base_urn", THREATS_BASE_URN)),
    )
    threats_content.append(
        ("ref_id", "name", "description", "name[fr]", "description[fr]")
    )
    for row_number, item in enumerate(techniques, start=2):
        threats_content.append(
            (
                item.ref_id,
                item.name,
                item.description,
                f"=TRADUIRE(B{row_number},\"en\",\"fr\")",
                f"=TRADUIRE(C{row_number},\"en\",\"fr\")",
            )
        )
    style_content_sheet(threats_content, (14, 48, 100, 48, 100))

    append_meta_rows(
        mitigations_meta,
        (("type", "reference_controls"), ("base_urn", MITIGATIONS_BASE_URN)),
    )
    mitigations_content.append(
        (
            "ref_id",
            "name",
            "csf_function",
            "category",
            "description",
            "name[fr]",
            "description[fr]",
        )
    )

    missing_csf_mappings: list[str] = []
    for row_number, item in enumerate(mitigations, start=2):
        csf_function = MITIGATION_CSF_FUNCTIONS.get(item.ref_id)
        if csf_function is None:
            csf_function = "protect"
            missing_csf_mappings.append(item.ref_id)
        mitigations_content.append(
            (
                item.ref_id,
                item.name,
                csf_function,
                "technical",
                item.description,
                f"=TRADUIRE(B{row_number},\"en\",\"fr\")",
                f"=TRADUIRE(E{row_number},\"en\",\"fr\")",
            )
        )
    style_content_sheet(
        mitigations_content,
        (14, 48, 18, 18, 100, 48, 100),
    )

    if missing_csf_mappings:
        refs = ", ".join(missing_csf_mappings)
        print(
            "⚠️  [WARNING] No preserved CSF function for "
            f"{refs}; defaulted to 'protect'.",
            file=sys.stderr,
        )

    temporary_output = output.with_name(f"{output.stem}.tmp.xlsx")
    temporary_output.unlink(missing_ok=True)
    try:
        workbook.save(temporary_output)
        temporary_output.replace(output)
    finally:
        workbook.close()
        temporary_output.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Cleanup and entry point
# ---------------------------------------------------------------------------


def cleanup(paths: Iterable[Path]) -> None:
    for path in paths:
        if path.exists():
            path.unlink()
            print(f'🗑️  [INFO] Deleted intermediate file: "{display_path(path)}"')


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Download MITRE ATT&CK Enterprise data and build the versioned "
            "CISO Assistant Excel library."
        )
    )
    parser.add_argument(
        "-k",
        "--keep",
        action="store_true",
        help="keep downloaded sources and intermediate Excel workbooks",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    generated_intermediates: list[Path] = []

    try:
        with create_session() as session:
            print_step_banner(1, "Download MITRE source files")
            generated_intermediates.extend(download_sources(session))

            print_step_banner(2, "Read ATT&CK release metadata")
            version = get_attack_version(session)
            readme = README_PATH.read_text(encoding="utf-8")
            license_text = LICENSE_PATH.read_text(encoding="utf-8")
            description = build_library_description(readme)
            copyright_text = extract_attack_license(license_text)
            print(f"✅ [OK] Detected ATT&CK Enterprise version: v{version}")

        print_step_banner(3, "Extract techniques and mitigations")
        techniques, mitigations = extract_attack_records(ENTERPRISE_ATTACK_PATH)
        print(f"✅ [OK] Retrieved {len(techniques)} ATT&CK techniques.")
        print(f"✅ [OK] Retrieved {len(mitigations)} ATT&CK mitigations.")

        print_step_banner(4, "Build intermediate Excel workbooks")
        generated_intermediates.extend(
            write_intermediate_workbooks(techniques, mitigations)
        )
        print(f'✅ [OK] Created: "{display_path(TECHNIQUES_PATH)}"')
        print(f'✅ [OK] Created: "{display_path(MEASURES_PATH)}"')

        print_step_banner(5, "Build final CISO Assistant workbook")
        output = SCRIPT_DIR / f"mitre-attack-v{version}.xlsx"
        build_final_workbook(
            output=output,
            version=version,
            description=description,
            copyright_text=copyright_text,
            techniques=techniques,
            mitigations=mitigations,
        )
        print(f'✅ [OK] Created: "{display_path(output)}"')

        print_step_banner(6, "Summary")
        print(f"- ATT&CK version: v{version}")
        print(f'- Final workbook: "{display_path(output)}"')
        if args.keep:
            print("ℹ️  [NOTE] Downloaded and intermediate files were kept (--keep).")

    except KeyboardInterrupt:
        print("❌ [ERROR] Interrupted by user.", file=sys.stderr)
        raise SystemExit(130)
    except Exception as exc:
        print(f"❌ [ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
    finally:
        if not args.keep:
            cleanup(generated_intermediates)


if __name__ == "__main__":
    main()
